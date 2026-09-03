"""API contract: the routes the frontend depends on, plus privacy operations."""

from __future__ import annotations

import asyncio
import csv
import io
import json
import re

import pytest
from fastapi.testclient import TestClient

from app.corpus.demo_profile import DEMO_PROFILE


@pytest.fixture
def client(tmp_path, monkeypatch, corpus_dir):
    """A client backed by a throwaway database and the bundled corpus."""
    from app.config import Settings, get_settings

    settings = Settings(
        demo_mode=True,
        database_url=f"sqlite:///{tmp_path / 'api.db'}",
        cache_dir=tmp_path / "cache",
        export_dir=tmp_path / "exports",
        corpus_dir=corpus_dir,
        fetch_delay_seconds=0.0,
        enable_browser_tier=False,
    )
    settings.ensure_dirs()
    get_settings.cache_clear()
    monkeypatch.setattr("app.config.get_settings", lambda: settings)

    import app.db as db_module

    engine = db_module.create_engine(
        settings.database_url, connect_args={"check_same_thread": False}
    )
    monkeypatch.setattr(db_module, "engine", engine)
    monkeypatch.setattr(
        db_module, "SessionLocal", db_module.sessionmaker(bind=engine, future=True)
    )
    # The schema is owned by Alembic now, so the test database is migrated the
    # same way production is rather than conjured with create_all().
    db_module.migrate_to_head(settings.database_url)

    from app.main import app

    with TestClient(app) as c:
        yield c
    get_settings.cache_clear()


async def drain_queue(limit: int = 10) -> int:
    """Run queued jobs to completion, as the worker process would.

    The API only enqueues now; a separate worker consumes. Tests drive that
    worker directly instead of pretending the API still does the work.
    """
    from app.jobs.worker import Worker

    worker = Worker()
    executed = 0
    for _ in range(limit):
        job_id = worker.claim_one()
        if job_id is None:
            break
        await worker.execute(job_id)
        executed += 1
    return executed


@pytest.fixture
def finished_run(client):
    profile = client.post("/api/profiles", json=DEMO_PROFILE.model_dump(mode="json")).json()
    queued = client.post(
        "/api/runs", json={"profile_id": profile["id"], "demo_mode": True}
    ).json()
    assert queued["stage"] == "queued"
    assert asyncio.run(drain_queue()) == 1, "the API must have enqueued exactly one job"

    state = client.get(f"/api/runs/{queued['id']}").json()
    assert state["stage"] == "awaiting_user_decision", state.get("errors")
    return profile, state


class TestMeta:
    def test_health_reports_the_mode_it_is_running_in(self, client):
        body = client.get("/api/health").json()
        assert body["status"] == "ok"
        assert body["demo_mode"] is True

    def test_capabilities_state_the_guarantees_and_the_limits(self, client):
        body = client.get("/api/capabilities").json()
        assert any("robots.txt" in g for g in body["guarantees"])
        assert any("never" in g and "outbound URL" in g for g in body["guarantees"])
        assert any("cannot predict" in limit for limit in body["limits"])

    def test_the_vocabulary_endpoint_exposes_every_controlled_list(self, client):
        vocab = client.get("/api/vocabulary").json()
        assert "FULL_RIDE_CONFIRMED" in vocab["funding_classification"]
        assert "NEEDS_OFFICIAL_CLARIFICATION" in vocab["eligibility"]


class TestProfiles:
    def test_a_profile_round_trips(self, client):
        created = client.post("/api/profiles", json=DEMO_PROFILE.model_dump(mode="json"))
        assert created.status_code == 201
        fetched = client.get(f"/api/profiles/{created.json()['id']}")
        assert fetched.json()["context"]["citizenship"] == "Kazakhstan"

    def test_validation_explains_the_consequence_of_each_gap(self, client):
        payload = DEMO_PROFILE.model_dump(mode="json")
        payload["academics"]["ielts"]["overall"] = None
        report = client.post("/api/profiles/validate", json=payload).json()
        english = next(g for g in report["gaps"] if "ielts" in g["field_path"])
        assert "hard requirement" in english["impact"]

    def test_a_missing_field_of_study_blocks_research(self, client):
        payload = DEMO_PROFILE.model_dump(mode="json")
        payload["context"]["intended_fields"] = []
        report = client.post("/api/profiles/validate", json=payload).json()
        assert report["can_proceed"] is False
        assert report["blocking_count"] == 1

    def test_an_invalid_ielts_band_is_rejected(self, client):
        payload = DEMO_PROFILE.model_dump(mode="json")
        payload["academics"]["ielts"]["overall"] = 6.2
        assert client.post("/api/profiles", json=payload).status_code == 422

    def test_a_grade_conversion_is_offered_with_its_caveat_never_applied(self, client):
        body = client.get("/api/profiles/conversions/methods?scale_label=KZ 5-point").json()
        assert body["methods"]
        assert "not a credential evaluation" in body["methods"][0]["source"]
        assert "never applied automatically" in body["note"].replace("is ever", "is never")


class TestResearchAndResults:
    def test_a_run_produces_results_with_evidence(self, client, finished_run):
        _, run = finished_run
        assert run["results_count"] >= 15
        assert run["claims_recorded"] > 50

        results = client.get(f"/api/runs/{run['id']}/results").json()
        assert all(r["claims"] for r in results if r["source_urls"])

    def test_the_summary_flags_demo_data(self, client, finished_run):
        _, run = finished_run
        summary = client.get(f"/api/runs/{run['id']}/summary").json()
        assert summary["demo_data"] is True
        assert summary["with_conflicts"] >= 1

    def test_results_can_be_filtered(self, client, finished_run):
        _, run = finished_run
        met = client.get(f"/api/runs/{run['id']}/results?eligibility=MET").json()
        assert met and all(r["eligibility"] == "MET" for r in met)

    def test_open_questions_are_exposed_for_the_user_to_chase(self, client, finished_run):
        _, run = finished_run
        questions = client.get(f"/api/runs/{run['id']}/questions").json()
        assert questions
        assert any(q["topic"] == "source conflict" for q in questions)

    def test_a_decision_is_recorded_with_its_reason(self, client, finished_run):
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        updated = client.post(
            f"/api/runs/{run['id']}/results/{result['id']}/decision",
            json={"decision": "approved", "reason": "best funded", "notes": "ask about housing"},
        ).json()
        assert updated["user_decision"] == "approved"
        assert updated["user_decision_reason"] == "best funded"

    def test_a_rejected_row_is_kept_not_deleted(self, client, finished_run):
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        client.post(
            f"/api/runs/{run['id']}/results/{result['id']}/decision",
            json={"decision": "rejected", "reason": "too expensive", "notes": ""},
        )
        rejected = client.get(f"/api/runs/{run['id']}/results?decision=rejected").json()
        assert len(rejected) == 1
        assert rejected[0]["user_decision_reason"] == "too expensive"

    def test_documents_cannot_be_collected_before_anything_is_approved(self, client, finished_run):
        _, run = finished_run
        response = client.post(f"/api/runs/{run['id']}/collect-documents")
        assert response.status_code == 400
        assert "Approve at least one" in response.json()["detail"]

    def test_documents_are_collected_for_approved_rows(self, client, finished_run):
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        client.post(f"/api/runs/{run['id']}/results/{result['id']}/decision",
                    json={"decision": "approved", "reason": "", "notes": ""})
        assert client.post(f"/api/runs/{run['id']}/collect-documents").status_code == 202
        assert asyncio.run(drain_queue()) == 1

        state = client.get(f"/api/runs/{run['id']}").json()
        assert state["stage"] == "completed"

        updated = client.get(f"/api/runs/{run['id']}/results/{result['id']}").json()
        assert updated["checklist"] is not None
        assert updated["checklist"]["ordered_steps"]


class TestStartingTwice:
    """One profile, one research at a time.

    idempotency_key was f"research:{run.id}" - unique per request by
    construction, so it deduplicated nothing. Two clicks produced two runs,
    both burning worker slots and outbound traffic.
    """

    def test_a_second_start_joins_the_run_already_in_flight(self, client):
        profile = client.post("/api/profiles", json=DEMO_PROFILE.model_dump(mode="json")).json()
        first = client.post("/api/runs", json={"profile_id": profile["id"], "demo_mode": True})
        assert first.status_code == 202

        second = client.post("/api/runs", json={"profile_id": profile["id"], "demo_mode": True})
        assert second.status_code == 409, second.text
        assert first.json()["id"] in second.json()["detail"]
        assert len(client.get("/api/runs").json()) == 1

    def test_an_idempotency_key_returns_the_same_run_rather_than_an_error(self, client):
        profile = client.post("/api/profiles", json=DEMO_PROFILE.model_dump(mode="json")).json()
        body = {"profile_id": profile["id"], "demo_mode": True}
        headers = {"Idempotency-Key": "the-same-click"}

        first = client.post("/api/runs", json=body, headers=headers)
        second = client.post("/api/runs", json=body, headers=headers)

        assert first.status_code == 202
        assert second.status_code == 202, second.text
        assert first.json()["id"] == second.json()["id"]
        assert len(client.get("/api/runs").json()) == 1

    def test_a_new_run_is_allowed_once_the_previous_one_is_finished(self, client, finished_run):
        profile, _ = finished_run
        again = client.post("/api/runs", json={"profile_id": profile["id"], "demo_mode": True})
        assert again.status_code == 202
        assert len(client.get("/api/runs").json()) == 2


class TestDocumentIdempotency:
    """Collecting documents twice must follow the shortlist, not its size.

    The audited defect keyed the job on the *count* of approved rows, so
    swapping one approval for another left the count unchanged, returned the
    old succeeded job, and the newly approved programme never got a checklist.
    """

    def test_swapping_an_approval_collects_documents_for_the_new_row(self, client, finished_run):
        _, run = finished_run
        rows = client.get(f"/api/runs/{run['id']}/results").json()
        first, second, third, replacement = rows[0], rows[1], rows[2], rows[3]

        def decide(result_id: str, decision: str) -> None:
            response = client.post(
                f"/api/runs/{run['id']}/results/{result_id}/decision",
                json={"decision": decision, "reason": "", "notes": ""},
            )
            assert response.status_code == 200, response.text

        for row in (first, second, third):
            decide(row["id"], "approved")
        assert client.post(f"/api/runs/{run['id']}/collect-documents").status_code == 202
        assert asyncio.run(drain_queue()) == 1

        # Same count, different set: one approval withdrawn, another added.
        decide(third["id"], "undecided")
        decide(replacement["id"], "approved")
        assert client.post(f"/api/runs/{run['id']}/collect-documents").status_code == 202
        assert asyncio.run(drain_queue()) == 1, "the second request must enqueue real work"

        added = client.get(f"/api/runs/{run['id']}/results/{replacement['id']}").json()
        assert added["checklist"] is not None
        assert added["checklist"]["ordered_steps"]

    def test_collecting_twice_for_the_same_shortlist_does_no_work_twice(
        self, client, finished_run
    ):
        _, run = finished_run
        row = client.get(f"/api/runs/{run['id']}/results").json()[0]
        client.post(
            f"/api/runs/{run['id']}/results/{row['id']}/decision",
            json={"decision": "approved", "reason": "", "notes": ""},
        )
        assert client.post(f"/api/runs/{run['id']}/collect-documents").status_code == 202
        assert asyncio.run(drain_queue()) == 1

        assert client.post(f"/api/runs/{run['id']}/collect-documents").status_code == 202
        assert asyncio.run(drain_queue()) == 0, "an unchanged shortlist must not re-fetch pages"


class TestRetry:
    """Retry must never be a way to lose work.

    The audited defect: retry deleted every result row but only reset stages
    that had failed. A successful run has every stage `done`, the runner skips
    all of them, and the user is left with an empty shortlist.
    """

    def test_retrying_a_finished_run_keeps_its_results_and_decisions(self, client, finished_run):
        _, run = finished_run
        before = client.get(f"/api/runs/{run['id']}/results").json()
        assert len(before) == 20
        kept = before[0]
        client.post(
            f"/api/runs/{run['id']}/results/{kept['id']}/decision",
            json={"decision": "approved", "reason": "shortlisted", "notes": "call the office"},
        )

        assert client.post(f"/api/runs/{run['id']}/retry").status_code == 200
        assert asyncio.run(drain_queue()) == 1

        state = client.get(f"/api/runs/{run['id']}").json()
        assert state["stage"] == "awaiting_user_decision", state.get("errors")
        after = client.get(f"/api/runs/{run['id']}/results").json()
        assert len(after) == len(before)

        decided = client.get(f"/api/runs/{run['id']}/results/{kept['id']}").json()
        assert decided["user_decision"] == "approved"
        assert decided["user_decision_reason"] == "shortlisted"
        assert decided["user_notes"] == "call the office"
        assert decided["decided_at"] is not None

    def test_a_full_retry_re_runs_every_stage_rather_than_skipping_them(self, client, finished_run):
        _, run = finished_run
        assert client.post(f"/api/runs/{run['id']}/retry").status_code == 200

        state = client.get(f"/api/runs/{run['id']}").json()
        assert {s["status"] for s in state["stages"]} == {"pending"}

    def test_retrying_one_stage_resets_it_and_everything_after_it(self, client, finished_run):
        _, run = finished_run
        response = client.post(f"/api/runs/{run['id']}/retry?stage=funding_discovery")
        assert response.status_code == 200

        stages = {s["stage"]: s["status"] for s in response.json()["stages"]}
        # Untouched: they ran before the named stage.
        assert stages["profile_validation"] == "done"
        assert stages["candidate_discovery"] == "done"
        assert stages["program_verification"] == "done"
        # Reset: the named stage and everything downstream of it.
        assert stages["funding_discovery"] == "pending"
        assert stages["assessment"] == "pending"

    def test_a_stage_retry_keeps_the_evidence_the_earlier_stages_produced(
        self, client, finished_run
    ):
        _, run = finished_run
        claims_before = len(client.get(f"/api/runs/{run['id']}/claims").json())
        assert claims_before > 0

        client.post(f"/api/runs/{run['id']}/retry?stage=assessment")
        assert asyncio.run(drain_queue()) == 1

        state = client.get(f"/api/runs/{run['id']}").json()
        assert state["stage"] == "awaiting_user_decision", state.get("errors")
        results = client.get(f"/api/runs/{run['id']}/results").json()
        assert len(results) == 20
        assert all(r["preference_score"] is not None for r in results), (
            "assessment must have scored the rows it kept, not an empty set"
        )


class TestExports:
    def test_csv_carries_the_disclaimer_sources_and_data_origin(self, client, finished_run):
        _, run = finished_run
        response = client.get(f"/api/runs/{run['id']}/export.csv")
        assert response.status_code == 200
        text = response.text
        assert "promises admission or an award" in text.splitlines()[0]

        rows = list(csv.DictReader(io.StringIO("\n".join(text.splitlines()[1:]))))
        assert rows
        assert all(r["Data origin"] for r in rows)
        assert any("DEMO FIXTURE" in r["Data origin"] for r in rows)
        assert all(r["Last verified date"] for r in rows)

    def test_csv_never_shows_an_unknown_as_a_blank(self, client, finished_run):
        _, run = finished_run
        text = client.get(f"/api/runs/{run['id']}/export.csv").text
        rows = list(csv.DictReader(io.StringIO("\n".join(text.splitlines()[1:]))))
        unknowns = [r for r in rows if r["Funding classification"] == "UNKNOWN"]
        assert unknowns
        for row in unknowns:
            assert row["Estimated funding gap"].startswith("not computable")
            assert row["Scholarship name"] in ("none found",) or row["Scholarship name"]

    def test_json_export_includes_every_claim(self, client, finished_run):
        _, run = finished_run
        body = json.loads(client.get(f"/api/runs/{run['id']}/export.json").text)
        assert "disclaimer" in body
        assert body["count"] == len(body["results"])
        assert any(r["claims"] for r in body["results"])

    def test_xlsx_has_shortlist_evidence_and_questions_sheets(self, client, finished_run):
        from openpyxl import load_workbook

        _, run = finished_run
        content = client.get(f"/api/runs/{run['id']}/export.xlsx").content
        workbook = load_workbook(io.BytesIO(content))
        assert workbook.sheetnames == ["Shortlist", "Evidence", "Open questions"]
        assert workbook["Evidence"].max_row > 20
        assert "promises admission or an award" in str(workbook["Shortlist"]["A1"].value)

    def test_an_unsupported_format_is_rejected(self, client, finished_run):
        _, run = finished_run
        assert client.get(f"/api/runs/{run['id']}/export.pdf").status_code == 400


class TestPrivacy:
    def test_deleting_a_profile_removes_every_run_and_result(self, client, finished_run):
        profile, run = finished_run
        assert client.delete(f"/api/profiles/{profile['id']}").status_code == 204
        assert client.get(f"/api/profiles/{profile['id']}").status_code == 404
        assert client.get("/api/runs").json() == []

    def test_a_profile_can_be_exported_in_full(self, client, finished_run):
        profile, _ = finished_run
        body = client.get(f"/api/profiles/{profile['id']}/export").json()
        assert body["profile"]["context"]["citizenship"] == "Kazakhstan"
        assert body["runs"]

    def test_the_audit_log_records_actions_without_applicant_data(self, client, finished_run):
        events = client.get("/api/audit").json()
        assert events
        assert "run_started" in {e["action"] for e in events}

        serialised = json.dumps(events)
        for phrase in ("Kazakhstan", "Demo Applicant", "IELTS", "TOEFL"):
            assert phrase not in serialised, f"audit log leaked {phrase!r}"

        # Bounded match: an unanchored "4.8" would hit the fractional seconds of
        # a timestamp and turn this into a test of the clock.
        assert not re.search(r"\b4\.8\b", serialised), "audit log leaked a GPA"

        # Detail payloads must hold counts and flags only, never free text
        # copied out of the profile.
        for event in events:
            for value in event["detail"].values():
                assert isinstance(value, int | float | bool) or (
                    isinstance(value, str) and len(value) < 60
                ), f"audit detail carries an unexpected payload: {value!r}"


class TestHealth:
    """A health check that cannot fail is not a health check.

    It reported the configuration only, so the probe stayed green with the
    database down and the platform kept routing traffic to a machine that
    could not answer one real request.
    """

    def test_a_healthy_service_reports_ok(self, client):
        body = client.get("/api/health").json()
        assert body["status"] == "ok"
        assert body["database"] == "ok"

    def test_an_unreachable_database_reports_degraded_with_503(self, client, monkeypatch):
        import sqlalchemy as sa

        import app.db as db_module

        monkeypatch.setattr(
            db_module, "engine", sa.create_engine("sqlite:///C:/nonexistent-dir/nope.db")
        )
        response = client.get("/api/health")

        assert response.status_code == 503
        body = response.json()
        assert body["status"] == "degraded"
        assert body["database"] == "unavailable"
        # The configuration fields survive, so the probe still says what build
        # it reached.
        assert "schema_version" in body


class TestRunViewJobReporting:
    """The run view reports the job the user is waiting on.

    A recheck job is queued months ahead. Reporting it as the run's current
    job made every screen believe work was in flight for ever: the collect
    button sat disabled at "Collecting…" and never came back.
    """

    def test_a_deferred_recheck_is_not_reported_as_work_in_flight(self, client, finished_run):
        _, run = finished_run
        state = client.get(f"/api/runs/{run['id']}").json()

        from app.db import SessionLocal
        from app.models import Job

        with SessionLocal() as session:
            assert session.query(Job).filter(
                Job.run_id == run["id"], Job.kind == "recheck"
            ).count() == 1, "the recheck must exist for this test to mean anything"

        assert state["job_status"] == "succeeded", "the research job is the one that matters"
        assert state["job_running"] is False
        assert state["next_recheck_at"] is not None

    def test_documents_can_still_be_collected_after_a_recheck_is_queued(
        self, client, finished_run
    ):
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        client.post(
            f"/api/runs/{run['id']}/results/{result['id']}/decision",
            json={"decision": "approved", "reason": "", "notes": ""},
        )
        assert client.post(f"/api/runs/{run['id']}/collect-documents").status_code == 202

        state = client.get(f"/api/runs/{run['id']}").json()
        assert state["job_status"] == "queued", "the documents job is what is now outstanding"
        assert asyncio.run(drain_queue()) == 1

        done = client.get(f"/api/runs/{run['id']}").json()
        assert done["stage"] == "completed"
        assert done["job_status"] == "succeeded", "a future recheck must not read as running"


class TestNotes:
    def test_saving_a_note_does_not_decide_the_row(self, client, finished_run):
        """Editing a note re-sent the decision, stamping decided_at on a row
        the applicant had not decided anything about."""
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        assert result["user_decision"] == "undecided"

        saved = client.patch(
            f"/api/runs/{run['id']}/results/{result['id']}/notes",
            json={"notes": "ask about housing"},
        )
        assert saved.status_code == 200
        body = saved.json()
        assert body["user_notes"] == "ask about housing"
        assert body["user_decision"] == "undecided"
        assert body["decided_at"] is None, "a note is not a decision"

    def test_a_note_survives_a_later_decision(self, client, finished_run):
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        client.patch(
            f"/api/runs/{run['id']}/results/{result['id']}/notes", json={"notes": "keep me"}
        )
        decided = client.post(
            f"/api/runs/{run['id']}/results/{result['id']}/decision",
            json={"decision": "approved", "reason": "", "notes": ""},
        ).json()
        assert decided["user_notes"] == "keep me"

    def test_the_notes_route_is_tenant_scoped_like_every_other(self, client, finished_run):
        _, run = finished_run
        result = client.get(f"/api/runs/{run['id']}/results").json()[0]
        response = client.patch(
            f"/api/runs/{'0' * 32}/results/{result['id']}/notes", json={"notes": "x"}
        )
        assert response.status_code == 404
