"""Exports.

The exported table is the one from section 7 of the specification, in full. It
carries the source links, the verification date and the demo-data flag, so an
exported file is as auditable as the screen it came from - a spreadsheet that
drops the provenance would be the easiest place for an unverified number to
start looking official.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import UTC, datetime
from typing import Any

from app.schemas.result import ProgramResult

COLUMNS: list[tuple[str, str]] = [
    ("university", "University"),
    ("country", "Country"),
    ("city", "City"),
    ("program", "Exact program"),
    ("degree", "Degree"),
    ("intake", "Intake"),
    ("ranking", "Ranking source, year and position"),
    ("eligibility", "Eligibility status"),
    ("admissions_fit", "Admissions fit"),
    ("applicant_scores", "Applicant scores"),
    ("published_requirements", "Published requirements"),
    ("missing_prerequisites", "Missing prerequisites"),
    ("scholarship_name", "Scholarship name"),
    ("scholarship_type", "Scholarship type"),
    ("funding_classification", "Funding classification"),
    ("scholarship_amount", "Scholarship amount and currency"),
    ("covered_costs", "Covered costs"),
    ("annual_cost", "Estimated annual cost of attendance"),
    ("funding_gap", "Estimated funding gap"),
    ("international_eligible", "International student eligibility"),
    ("application_mode", "Automatic or separate application"),
    ("admission_deadline", "Admission deadline"),
    ("scholarship_deadline", "Scholarship deadline"),
    ("renewal_requirements", "Renewal requirements"),
    ("climate_fit", "Climate fit"),
    ("city_fit", "City fit"),
    ("workload_fit", "Workload fit"),
    ("career_notes", "Career/internship opportunities"),
    ("post_study_work", "Post-study work information"),
    ("admission_documents", "Required admission documents"),
    ("scholarship_documents", "Required scholarship documents"),
    ("conflicts", "Conflicts or uncertainties"),
    ("last_verified", "Last verified date"),
    ("source_links", "Official source links"),
    ("preference_score", "Preference score (not a probability)"),
    ("user_status", "User status"),
    ("user_notes", "User notes"),
    ("data_origin", "Data origin"),
]

DISCLAIMER = (
    "ASHYQ Apply export. Eligibility, admissions fit and funding fit describe published criteria "
    "only. Nothing here predicts or promises admission or an award. Verify every value against "
    "the official source link before acting on it."
)


def _row(result: ProgramResult) -> dict[str, Any]:
    best = _best_scholarship(result)
    gap = result.funding_gap
    a = result.preference_score

    return {
        "university": result.university,
        "country": result.country,
        "city": result.city,
        "program": result.program,
        "degree": result.degree.value,
        "intake": result.intake,
        "ranking": "; ".join(f"{r.source} {r.year}: {r.position}" for r in result.rankings) or "not found",
        "eligibility": result.eligibility.value,
        "admissions_fit": result.admissions_fit.value,
        "applicant_scores": "; ".join(
            f"{c.requirement}={c.applicant_value}"
            for c in result.requirement_checks
            if c.applicant_value is not None
        ),
        "published_requirements": "; ".join(
            f"{c.requirement}: {c.published_value} [{c.status.value}]"
            for c in result.requirement_checks
            if c.published_value is not None
        ),
        "missing_prerequisites": "; ".join(result.missing_prerequisites) or "none identified",
        "scholarship_name": best.name if best else "none found",
        "scholarship_type": best.scholarship_type.value if best else "",
        "funding_classification": result.best_funding_classification.value,
        "scholarship_amount": (
            str(best.amount) if best and best.amount
            else (f"{best.amount_is_percentage_of_tuition:g}% of tuition"
                  if best and best.amount_is_percentage_of_tuition else "not published")
        ),
        "covered_costs": "; ".join(
            f"{c.category.value}={c.covered}" for c in (best.coverage if best else [])
        ) or "not published",
        "annual_cost": str(gap.total_cost) if gap and gap.total_cost else "not found",
        "funding_gap": (
            str(gap.gap) if gap and gap.computable and gap.gap
            else f"not computable — {gap.reason}" if gap else "not computed"
        ),
        "international_eligible": best.international_eligible if best else "unknown",
        "application_mode": best.application_mode.value if best else "unknown",
        "admission_deadline": (
            f"{result.admission_deadline.isoformat()} {result.admission_deadline_timezone or ''}".strip()
            if result.admission_deadline else "not found"
        ) + (" (PASSED)" if result.deadline_passed else ""),
        "scholarship_deadline": (
            f"{best.deadline.isoformat()} {best.deadline_timezone or ''}".strip()
            if best and best.deadline else "not found"
        ),
        "renewal_requirements": "; ".join(best.renewal_requirements) if best else "",
        "climate_fit": result.climate_fit,
        "city_fit": result.city_fit,
        "workload_fit": result.workload_fit,
        "career_notes": result.career_notes,
        "post_study_work": result.post_study_work,
        "admission_documents": "; ".join(
            d.name for d in (result.checklist.admission_documents if result.checklist else [])
        ) or "not collected (approve to collect)",
        "scholarship_documents": "; ".join(
            d.name for d in (result.checklist.scholarship_documents if result.checklist else [])
        ) or "not collected (approve to collect)",
        "conflicts": "; ".join(
            f"{c.subject}: {c.values}" for c in result.conflicts
        ) or ("; ".join(q.question for q in result.unresolved) or "none"),
        "last_verified": result.last_verified.isoformat() if result.last_verified else "never",
        "source_links": " | ".join(result.source_urls),
        "preference_score": f"{a.total:.2f} / {a.max_possible:.2f}" if a else "",
        "user_status": result.user_decision.value,
        "user_notes": result.user_notes,
        "data_origin": (
            "DEMO FIXTURE (synthetic)"
            if any(u.startswith("fixture://") for u in result.source_urls)
            else "live official sources"
        ),
    }


def _best_scholarship(result: ProgramResult):
    if not result.scholarships:
        return None
    order = ["FULL_RIDE_CONFIRMED", "FULL_TUITION", "LARGE_GRANT", "PARTIAL",
             "NEED_BASED_POSSIBLE", "UNKNOWN", "NOT_ELIGIBLE"]
    return min(result.scholarships, key=lambda s: order.index(s.classification.value))


#: Characters that make a spreadsheet read a cell as a formula rather than as
#: text. Tab and carriage return are here because Excel strips leading
#: whitespace before deciding, so "\t=1+1" is still a formula.
_FORMULA_LEADERS = ("=", "+", "-", "@", "\t", "\r")
#: A plain negative number is not a formula, and mangling one would corrupt a
#: funding gap with the very defence meant to protect it.
_PLAIN_NUMBER = re.compile(r"^-\d+(?:[.,]\d+)*$")


def neutralise_formula(value: str) -> str:
    """Stop a cell being executed, without changing what it says.

    Every string in an export came off a university web page, which is
    untrusted input. A scholarship named `=HYPERLINK("http://attacker","Click")`
    executes when the file is opened. Prefixing an apostrophe is the
    spreadsheet convention for "treat the rest as text": Excel, LibreOffice and
    Numbers all honour it, and it is not shown to the reader.
    """
    if not isinstance(value, str) or not value:
        return value
    if _PLAIN_NUMBER.match(value.strip()):
        return value
    if value.startswith(_FORMULA_LEADERS):
        return "'" + value
    return value


def _safe_row(result: ProgramResult) -> dict[str, Any]:
    """One exported row with every cell neutralised."""
    return {
        key: neutralise_formula(value) if isinstance(value, str) else value
        for key, value in _row(result).items()
    }


def to_csv(results: list[ProgramResult]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    writer.writerow([f"# {DISCLAIMER}"])
    writer.writerow([label for _, label in COLUMNS])
    for r in results:
        row = _safe_row(r)
        writer.writerow([row[key] for key, _ in COLUMNS])
    return buf.getvalue()


def to_json(results: list[ProgramResult], meta: dict | None = None) -> str:
    return json.dumps(
        {
            "disclaimer": DISCLAIMER,
            "exported_at": datetime.now(UTC).isoformat(),
            "meta": meta or {},
            "count": len(results),
            "results": [r.model_dump(mode="json") for r in results],
        },
        indent=2,
        ensure_ascii=False,
    )


def to_xlsx(results: list[ProgramResult], meta: dict | None = None) -> bytes:
    """Three sheets: the shortlist, every claim, and every open question."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlist"

    ws.append([DISCLAIMER])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    warn = ws.cell(row=1, column=1)
    warn.font = Font(bold=True, color="7F4F00")
    warn.fill = PatternFill("solid", fgColor="FFF3CD")
    warn.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 32

    ws.append([label for _, label in COLUMNS])
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    status_fills = {
        "FULL_RIDE_CONFIRMED": "C6EFCE", "FULL_TUITION": "DDEBF7", "LARGE_GRANT": "FFF2CC",
        "NOT_ELIGIBLE": "F8CBAD", "UNKNOWN": "EDEDED",
    }
    for r in results:
        row = _safe_row(r)
        ws.append([row[key] for key, _ in COLUMNS])
        idx = ws.max_row
        fill = status_fills.get(row["funding_classification"])
        if fill:
            ws.cell(row=idx, column=[k for k, _ in COLUMNS].index("funding_classification") + 1).fill = (
                PatternFill("solid", fgColor=fill)
            )

    widths = {"university": 30, "program": 34, "published_requirements": 46, "conflicts": 44,
              "source_links": 52, "post_study_work": 46, "covered_costs": 30}
    for i, (key, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(key, 20)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{max(2, ws.max_row)}"

    cs = wb.create_sheet("Evidence")
    cs.append(["University", "Program", "Claim type", "Value", "Status", "Specificity",
               "Accessed", "Source URL", "Excerpt"])
    for cell in cs[1]:
        cell.font = Font(bold=True)
    for r in results:
        for claim in r.claims:
            cs.append([
                r.university, r.program, claim.claim_type.value,
                str(claim.normalized_value)[:200], claim.status.value,
                claim.source_specificity.value, claim.accessed_at.isoformat(),
                claim.source_url, claim.original_text_excerpt[:400],
            ])
    for col, width in zip("ABCDEFGHI", (28, 30, 26, 30, 22, 22, 26, 52, 70), strict=False):
        cs.column_dimensions[col].width = width
    cs.freeze_panes = "A2"

    qs = wb.create_sheet("Open questions")
    qs.append(["University", "Program", "Topic", "Question", "Why it matters", "Blocking"])
    for cell in qs[1]:
        cell.font = Font(bold=True)
    for r in results:
        for conflict in r.conflicts:
            qs.append([r.university, r.program, "source conflict", conflict.subject,
                       f"Values seen: {conflict.values}", "yes"])
        for q in r.unresolved:
            qs.append([r.university, r.program, q.topic, q.question, q.why_it_matters,
                       "yes" if q.blocking else "no"])
    for col, width in zip("ABCDEF", (28, 30, 22, 56, 56, 10), strict=False):
        qs.column_dimensions[col].width = width
    qs.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
